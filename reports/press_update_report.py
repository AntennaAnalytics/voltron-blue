"""
Press Premium SVOD Update Report
Refactored version for Streamlit cloud environment
"""

import os
import datetime
import requests
import pandas as pd
import openpyxl
from openpyxl.utils import coordinate_to_tuple
import streamlit as st


def replace_decimal_separator(value):
    """Keep decimal point as period for numeric values (standard format)"""
    # Simply return the value without conversion - use period as decimal separator
    return value


# Query definitions
QUERIES = {
    "Sign-ups": {
        "query_id": 20491,
        "tab": "Sign-ups",
        "paste_cell": "A9",
        "date_cell": "B6",
        "pivot": True,
        "params": {
            "api_key": "QSTJ3AfkTmgvLmnlKPuLeSlSeYE36BBnTvjAMR7J",
            "p_A+ release date": "230913",
            "p_Latest Monthly Data": "2023-08-01"
        }
    },
    "Gross Adds": {
        "query_id": 20492,
        "tab": "Gross Adds",
        "paste_cell": "A10",
        "date_cell": "B7",
        "pivot": True,
        "params": {
            "api_key": "D7fcPON9HY9lfy9fcvU6oldWwHf65n1UYSCk1RWX",
            "p_A+ release date": "230913",
            "p_Latest Monthly Data": "2023-08-01"
        }
    },
    "Churn Weighted": {
        "query_id": 20494,
        "tab": "Churn",
        "paste_cell": "A10",
        "date_cell": None,
        "pivot": False,
        "params": {
            "api_key": "x8y07f11DW84Wf2aMExLBF0HWY2Ng12RCY6w4XEL",
            "p_A+ release date": "230913",
            "p_Latest Monthly Data": "2023-08-01"
        }
    },
    "Churn Monthly": {
        "query_id": 20493,
        "tab": "Churn",
        "paste_cell": "C10",
        "date_cell": "B6",
        "pivot": True,
        "include_row_names": False,
        "params": {
            "api_key": "WN7YMBY9xvC2SQnRPTIQpIu4EmCRYgwtXLcYg4a1",
            "p_A+ release date": "230913",
            "p_Latest Monthly Data": "2023-08-01"
        }
    }
}

REDASH_BASE_URL = "https://redash.uuid-p.antennaanalytics.com/api/queries/{query_id}/results.json"


@st.cache_data(ttl=3600)
def fetch_redash_query(query_id, params):
    """
    Fetch data from Redash API
    Cached for 1 hour to improve performance
    """
    api_url = REDASH_BASE_URL.format(query_id=query_id)

    try:
        response = requests.get(api_url, params=params, timeout=30)

        if response.status_code != 200:
            return None, f"API request failed with status {response.status_code}: {response.text}"

        json_data = response.json()
        query_result = json_data.get("query_result", {})
        data = query_result.get("data", {})
        rows = data.get("rows", [])
        columns = data.get("columns", [])

        return {"rows": rows, "columns": columns}, None

    except Exception as e:
        return None, f"Error fetching data: {str(e)}"


def process_query_data(query_name, query_info):
    """
    Process a single query: fetch and transform data
    Returns: (dataframe, error_message)
    """
    query_id = query_info["query_id"]
    query_params = query_info["params"]
    needs_pivot = query_info.get("pivot", False)
    include_row_names = query_info.get("include_row_names", True)

    # Fetch data from Redash
    result, error = fetch_redash_query(query_id, query_params)
    if error:
        return None, error

    rows = result["rows"]
    columns = result["columns"]
    col_names = [col.get("name") for col in columns]

    if not rows:
        return pd.DataFrame(), None

    # Create DataFrame
    df = pd.DataFrame(rows)

    if needs_pivot:
        # Pivot the data: month as rows, service as columns
        if len(col_names) >= 3:
            month_col = col_names[0]
            service_col = col_names[1]
            value_col = col_names[2]

            df_pivot = df.pivot(index=month_col, columns=service_col, values=value_col)

            # Sort month ascending and service columns alphabetically
            df_pivot = df_pivot.sort_index(ascending=True)
            df_pivot = df_pivot.reindex(sorted(df_pivot.columns), axis=1)

            # Reset index if we want to include month column
            if include_row_names:
                df_pivot = df_pivot.reset_index()

            return df_pivot, None
        else:
            return df, None
    else:
        return df, None


def generate_excel_file(query_data, output_path):
    """
    Generate Excel file from query data using openpyxl
    This replaces xlwings for cloud compatibility
    """
    try:
        # Load the template
        template_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            "templates",
            "Antenna for Press_PremiumSVOD_ yyyymmdd.xlsx"
        )

        if not os.path.exists(template_path):
            return False, f"Template file not found: {template_path}"

        # Load workbook
        wb = openpyxl.load_workbook(template_path)

        # Today's date
        today_date_str = datetime.date.today().strftime("%Y-%m-%d")

        # Process each query and update worksheets
        for query_name, query_info in QUERIES.items():
            tab_name = query_info["tab"]
            paste_cell = query_info["paste_cell"]
            date_cell = query_info["date_cell"]

            df = query_data.get(query_name)
            if df is None or df.empty:
                continue

            # Get worksheet
            if tab_name not in wb.sheetnames:
                continue

            ws = wb[tab_name]

            # Convert DataFrame to rows (without header)
            data_rows = df.values.tolist()

            # Parse the paste cell (e.g., "A9" -> row 9, col 1)
            row_start, col_start = coordinate_to_tuple(paste_cell)

            # Write data to worksheet
            for i, data_row in enumerate(data_rows):
                for j, value in enumerate(data_row):
                    cell = ws.cell(row=row_start + i, column=col_start + j)
                    # Write value directly - keep period as decimal separator
                    cell.value = value

            # Update date cell if specified
            if date_cell:
                ws[date_cell] = today_date_str

        # Save workbook
        wb.save(output_path)
        return True, None

    except Exception as e:
        return False, f"Error generating Excel file: {str(e)}"


def generate_press_update_report():
    """
    Main function to generate the Press Premium SVOD Update report
    Returns: dict with success status, data, and file path
    """
    result = {
        "success": False,
        "error": None,
        "details": None,
        "data": {
            "signups": None,
            "gross_adds": None,
            "churn": None
        },
        "data_summary": {
            "signups_rows": 0,
            "gross_adds_rows": 0,
            "churn_rows": 0
        },
        "excel_file": None,
        "report_date": datetime.date.today().strftime("%Y-%m-%d")
    }

    query_data = {}

    try:
        # Fetch and process all queries
        for query_name, query_info in QUERIES.items():
            df, error = process_query_data(query_name, query_info)

            if error:
                result["error"] = f"Error processing {query_name}"
                result["details"] = error
                return result

            query_data[query_name] = df

            # Store data for display
            if query_name == "Sign-ups":
                result["data"]["signups"] = df
                result["data_summary"]["signups_rows"] = len(df)
            elif query_name == "Gross Adds":
                result["data"]["gross_adds"] = df
                result["data_summary"]["gross_adds_rows"] = len(df)
            elif query_name in ["Churn Weighted", "Churn Monthly"]:
                # Combine churn data for display
                if result["data"]["churn"] is None:
                    result["data"]["churn"] = df
                result["data_summary"]["churn_rows"] += len(df)

        # Generate Excel file
        output_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            "Output"
        )
        os.makedirs(output_dir, exist_ok=True)

        current_date_str = datetime.date.today().strftime("%Y%m%d")
        output_filename = f"Antenna for Press_PremiumSVOD_{current_date_str}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        success, error = generate_excel_file(query_data, output_path)

        if not success:
            result["error"] = "Failed to generate Excel file"
            result["details"] = error
            return result

        result["excel_file"] = output_path
        result["success"] = True

        return result

    except Exception as e:
        result["error"] = "Unexpected error during report generation"
        result["details"] = str(e)
        return result
